# -*- coding: utf-8 -*-
"""
Created on Sat Oct  2 20:58:45 2021

@author: Andry
"""
import openpyxl, shutil
#______________________________________________________________________________
def Name_Uvolenogo (case_name, name_kadet, linum):#запись имени увольняемного
    excel_file = openpyxl.load_workbook(file_final) #открывает файл
    excel_list = excel_file[list_number[linum]]             #выбирает лист файла
    excel_list[case_name] = name_kadet             #перезапись ячейки
    excel_file.save(file_final)                     #сохранение изменений
   
def Data_Udolnenia (case_dara, period_uval, linum):#запись даты и времени конца увольнения
    excel_file = openpyxl.load_workbook(file_final) #открывает файл
    excel_list = excel_file[list_number[linum]]             #выбирает лист файла
    excel_list[case_dara] = period_uval             #перезапись ячейки
    excel_file.save(file_final)                     #сохранение изменений

def Data_Vudachi (case_write, vudacha, linum): #запись даты выдачи увольнительной записки
    excel_file = openpyxl.load_workbook(file_final) #открывает файл
    excel_list = excel_file[list_number[linum]]             #выбирает лист файла
    excel_list[case_write] = vudacha            #перезапись ячейки
    excel_file.save(file_final)                     #сохранение изменений
    
def Zanesenie_Dannuh (i, linum, name_kadet, period_uval, vudacha):
    '''
    Функция принимает значение номера ув.записки на странице для печати, 
    т.е. отвечает за изменение данных в ОДНОЙ ув.записке.
    где i - место записки на листе
    '''
    ghg = [['A3','A13','A23','A33','A43','F3','F13','F23','F33','F43'], #ячейки 
           ['A5','A15','A25','A35','A45','F5','F15','F25','F35','F45'], #для замены 
          ['C10','C20','C30','C40','C50','H10','H20','H30','H40','H50']]#данных
    case_name = ghg[0][i]           #имя
    case_dara = ghg[1][i]           #дата конца увольнения
    case_write = ghg[2][i]          #дата выдачи  
    Name_Uvolenogo(case_name, name_kadet, linum)
    Data_Udolnenia(case_dara, period_uval, linum)
    Data_Vudachi(case_write, vudacha, linum)
    
def Rabota_na_odnom_Liste (n, linum): #где n - кол-во эл-в на одной странице
    for i in range(0,n): #прогоняет по одной странице
        name_kadet = CADET_DATA[0][i]  
        period_uval = CADET_DATA[1][i]
        vudacha = CADET_DATA[2][i]
        Zanesenie_Dannuh (i, linum, name_kadet, period_uval, vudacha) 
        
def Rabota_nad_vsem_Failom (value_uvol): 
    '''
    Функция принимает значение кол-ва ув.записок для печати - value_uvol
    '''
    if value_uvol >10:
        a= (value_uvol // 10) #целое, кол-во полностью заполн. страниц
        b= (value_uvol % 10)   #остаток, на последней неполной странице
        if a > 10 : 
            print("ERROR")
            exit ()
            
        for c in range(a):#полные страницы
            Rabota_na_odnom_Liste(10, c)   
            
        for m in range(b):#неполная страница
            Rabota_na_odnom_Liste(b, a)
    else:
        for m in range(value_uvol):#неполная страница
            print(list_number[0])
            Rabota_na_odnom_Liste(value_uvol, 0)
#______________________________________________________________________________

global list_number, CADET_DATA, bar
file_initial = 'uvol_zap_initial.xlsx' #имя файла
file_final = 'uvol_zap_final.xlsx'
list_number = ['1','2','3','4','5','6','7','8','9','10']
shutil.copyfile("uvol_zap_initial.xlsx", "uvol_zap_final.xlsx")


# надо ебануть загрузку списка отмеченных(в увольнение) людей 
# и проставленных пользователем дат в этот массив
CADET_DATA = [ ["Косенко Евгений","Спирин Вячеслав","Чинакаев Дмитрий"],
               ["19-00 12.04.2021","20-00 27.08.2021","12-00 30.09.2021"],
               ["10 сентября 2021","10 сентября 2021","10 сентября 2021"] ]

            

#______________________________________________________________________________
Rabota_nad_vsem_Failom(3) # принимает кол-во увольнительных


