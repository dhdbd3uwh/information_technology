# -*- coding: utf-8 -*-
"""
Created on Tue Oct 19 12:43:18 2021

@author: Andry
"""

import openpyxl, sys
import xlwings as xw

from variables import list_number, file_final, CADET_DATA2, CADET_DATA1
from variables import schet_name, con, period_uval


# ______________________________________________________________________________
def Udalenie_Lishnih_Listov(value_uvol):
    print('\nУдаление ненужных листов . . .')
    '''
    Функция удаляет незаполненные листы, которые не нужны при печати, из конечного документа 
    linum - имя удаляемого листа(его номер в списке)
    Функция принимает кол-во увол.записок (value_uvol)
    '''
    value_list_for_print = 0  # кол-во листов для печати
    value_list = (value_uvol // 10)  # кол-во полностью заполн. страниц
    value_list_unfull = (value_uvol % 10)  # кол-во записок на последней неполной странице
    if value_list_unfull > 0 and value_list_unfull < 10:
        value_list_for_print = value_list + 1
    elif value_list_unfull == 0:
        value_list_for_print = value_list
    else:
        print("ERROR delete list in Udalenie_Lishnih_Listov()")
        sys.exit(1)
    print("листов на печать: ", value_list_for_print)

    # (кол-во удаляемых листов)=(кол-во всех листов)-(кол-во листов для печати)
    value_list_for_delete = len(list_number) - value_list_for_print
    print("лишних листов на удаление: ", value_list_for_delete)

    for i in range(value_list_for_print, (len(list_number))):
        linum = i  # номер листа для удаления
        excel_file = openpyxl.load_workbook(file_final)  # открывает файл
        excel_file.remove(excel_file[list_number[linum]])  # удаляем лист
        excel_file.save(file_final)


# ______________________________________________________________________________
def Schet_Lidey():
    '''
    Функция возвращает имя кадета по очереди списка, попутно меняя значение 
    шкалы состояния процесса работы программы
    '''
    global schet_name
    print('.')
    name_kadet = CADET_DATA2[schet_name]
    schet_name += 1
    return name_kadet


# ______________________________________________________________________________
def Setting_Variables():
    '''
    Задание необходимых переменных, списков и ввод данных с консоли 
    '''
    er = period_uval[1]
    vg = period_uval[0]
    return er, vg


# ______________________________________________________________________________
def Name_Uvolenogo(case_name, name_kadet, linum):  # запись имени увольняемного
    excel_file = openpyxl.load_workbook(file_final)  # открывает файл
    excel_list = excel_file[list_number[linum]]  # выбирает лист файла
    excel_list[case_name] = name_kadet  # перезапись ячейки
    excel_file.save(file_final)  # сохранение изменений


def Data_Udolnenia(case_dara, period_uval, linum):  # запись даты и времени конца увольнения
    excel_file = openpyxl.load_workbook(file_final)  # открывает файл
    excel_list = excel_file[list_number[linum]]  # выбирает лист файла
    excel_list[case_dara] = period_uval  # перезапись ячейки
    excel_file.save(file_final)  # сохранение изменений


def Data_Vudachi(case_write, vudacha, linum):  # запись даты выдачи увольнительной записки
    excel_file = openpyxl.load_workbook(file_final)  # открывает файл
    excel_list = excel_file[list_number[linum]]  # выбирает лист файла
    excel_list[case_write] = vudacha  # перезапись ячейки
    excel_file.save(file_final)  # сохранение изменений


# ______________________________________________________________________________
def Zanesenie_Dannuh(i, linum, name_kadet, per_uval, vuda):
    '''
    Функция принимает значение номера ув.записки на странице для печати, 
    т.е. отвечает за изменение данных в ОДНОЙ ув.записке.
    где i - место записки на листе
    '''
    ghg = [['A3', 'A13', 'A23', 'A33', 'A43', 'F3', 'F13', 'F23', 'F33', 'F43'],  # ячейки
           ['A5', 'A15', 'A25', 'A35', 'A45', 'F5', 'F15', 'F25', 'F35', 'F45'],  # для замены
           ['C10', 'C20', 'C30', 'C40', 'C50', 'H10', 'H20', 'H30', 'H40', 'H50']]  # данных
    case_name = ghg[0][i]  # имя
    case_dara = ghg[1][i]  # дата конца увольнения
    case_write = ghg[2][i]  # дата выдачи
    Name_Uvolenogo(case_name, name_kadet, linum)
    Data_Udolnenia(case_dara, per_uval, linum)
    Data_Vudachi(case_write, vuda, linum)


# ______________________________________________________________________________
def Rabota_na_odnom_Liste(n, linum):  # где n - кол-во эл-в на одной странице
    '''
    Функция принимает значение небходимого кол-ва ув.записок на одной странице 
    и номер листа в xlsx (linum)
    '''
    global Setting_Variables_list
    for i in range(n):  # прогоняет по одной странице
        name_kadet_iz = Schet_Lidey()
        Zanesenie_Dannuh(i, linum, name_kadet_iz, Setting_Variables_list[0], Setting_Variables_list[1])


# ______________________________________________________________________________
def Rabota_nad_vsem_Failom(value_uvol):
    '''
    Функция принимает значение кол-ва ув.записок для печати - value_uvol
    и организует заполнение данных в документе при помощи вызова ф-и 
    "работы на одном листе" необходимое кол-во раз и с небходимыми параметрами
    '''
    global Setting_Variables_list
    Setting_Variables_list = Setting_Variables()

    if value_uvol > 10:
        a = (value_uvol // 10)  # целое, кол-во полностью заполн. страниц
        b = (value_uvol % 10)  # остаток, на последней неполной странице
        if a > 10:
            print("ERROR full list in Rabota_nad_vsem_Failom()")
            sys.exit(1)

        for c in range(a):  # полные страницы
            Rabota_na_odnom_Liste(10, c)

        for m in range(b):  # неполная страница
            Rabota_na_odnom_Liste(b, a)
    elif value_uvol <= 10:
        Rabota_na_odnom_Liste(value_uvol, 0)
    # после занесения данных удаляем лишние листы
    Udalenie_Lishnih_Listov(value_uvol)


# ______________________________________________________________________________

def Print_File_Final(value_uvol):
    '''
    Функция отправляет итоговый фоайл с заполнеными увол. записками в очередь 
    печати принтера, установленого по умолчанию
    '''
    wb = openpyxl.load_workbook(file_final)
    res = len(wb.sheetnames)
    wb = xw.Book(file_final)
    sh2 = wb.sheets
    sh2.api.PrintOut(From=1, To=res, Copies=1)

def Start_Filling_Process (value_uvol):
    Rabota_nad_vsem_Failom(value_uvol)
    Print_File_Final(value_uvol)

# ______________________________________________________________________________


'''
ФУНКЦИИ РАБОТЫ С БД    
'''

# ______________________________________________________________________________
def BD_Cadet_Data():
    CADET_DATA1.clear()
    CADET_DATA2.clear()
    #with con:
    cur = con.cursor()
    cur.execute("SELECT * FROM cadet ORDER BY name")  # имя таблицы
    rows = cur.fetchall()
    for row in rows:  # заполнение списка именами
        # если фамилия и имя в одном столбце
        CADET_DATA1.append(row[1])

    #with con:
    cur.execute("SELECT * FROM cadet_print ORDER BY name")  # имя таблицы
    rows = cur.fetchall()
    for row in rows:  # заполнение списка именами
        # если фамилия и имя в одном столбце
        CADET_DATA2.append(row[1])


def Insert_SQL_Table(x):
    '''
    Функция добаляет в таблицу имя Х
    '''
    #with con:
    cur = con.cursor()
    sql = f"INSERT INTO cadet_print(name) VALUES ('{x}')"
    cur.execute(sql)
    con.commit()

def Insert_SQL_Table1(x):
    '''
    Функция добаляет в таблицу имя Х
    '''
    #with con:
    cur = con.cursor()
    sql = f"INSERT INTO cadet(name) VALUES ('{x}')"
    cur.execute(sql)
    con.commit()

def Delete_Odnogo_SQL_Table(x):
    '''
    Функция удаляет челоека
    '''
    cur = con.cursor()
    sql = f"DELETE FROM cadet_print WHERE name='{x}'"
    cur.execute(sql)
    con.commit()

def Delete_Odnogo_SQL_Table1(x):
    '''
    Функция удаляет челоека
    '''
    cur = con.cursor()
    sql = f"DELETE FROM cadet WHERE name='{x}'"
    cur.execute(sql)
    con.commit()

def Clean_Table():
    cur = con.cursor()
    sql = "TRUNCATE TABLE cadet"
    cur.execute(sql)
    con.commit()

    cur = con.cursor()
    sql = "TRUNCATE TABLE cadet_print"
    cur.execute(sql)
    con.commit()


